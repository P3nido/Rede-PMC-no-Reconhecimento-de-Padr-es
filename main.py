import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# Conjuntos de dados (embutidos como vetores)
#
# Cada linha = [x1, x2, x3, x4, d1, d2, d3]
#   x1 - teor de agua
#   x2 - grau de acidez
#   x3 - temperatura
#   x4 - tensao superficial
#   d1, d2, d3 - saida desejada (codificacao one-hot do conservante A, B ou C)
#
# Tipo A -> [1, 0, 0]   Tipo B -> [0, 1, 0]   Tipo C -> [0, 0, 1]
# =============================================================================

# Subconjunto de TREINAMENTO (130 amostras)
dados_treinamento = np.array([
    [0.3841, 0.2021, 0.0000, 0.2438, 1.0000, 0.0000, 0.0000], [0.1765, 0.1613, 0.3401, 0.0843, 1.0000, 0.0000, 0.0000], [0.3170, 0.5786, 0.3387, 0.4192, 0.0000, 1.0000, 0.0000],
    [0.2467, 0.0337, 0.2699, 0.3454, 1.0000, 0.0000, 0.0000], [0.6102, 0.8192, 0.4679, 0.4762, 0.0000, 1.0000, 0.0000], [0.7030, 0.7784, 0.7482, 0.6562, 0.0000, 0.0000, 1.0000],
    [0.4767, 0.4348, 0.4852, 0.3640, 0.0000, 1.0000, 0.0000], [0.7589, 0.8256, 0.6514, 0.6143, 0.0000, 0.0000, 1.0000], [0.1579, 0.3641, 0.2551, 0.2919, 1.0000, 0.0000, 0.0000],
    [0.5561, 0.5602, 0.5605, 0.2105, 0.0000, 1.0000, 0.0000], [0.3267, 0.2974, 0.0343, 0.1466, 1.0000, 0.0000, 0.0000], [0.2303, 0.0942, 0.3889, 0.1713, 1.0000, 0.0000, 0.0000],
    [0.2953, 0.2963, 0.2600, 0.3039, 1.0000, 0.0000, 0.0000], [0.5797, 0.4789, 0.5780, 0.3048, 0.0000, 1.0000, 0.0000], [0.5860, 0.5250, 0.4792, 0.4021, 0.0000, 1.0000, 0.0000],
    [0.7045, 0.6933, 0.6449, 0.6623, 0.0000, 0.0000, 1.0000], [0.9134, 0.9412, 0.6078, 0.5934, 0.0000, 0.0000, 1.0000], [0.2333, 0.4943, 0.2525, 0.2567, 1.0000, 0.0000, 0.0000],
    [0.2676, 0.4172, 0.2775, 0.2721, 1.0000, 0.0000, 0.0000], [0.4850, 0.5506, 0.5269, 0.6036, 0.0000, 1.0000, 0.0000], [0.2434, 0.2567, 0.2312, 0.2624, 1.0000, 0.0000, 0.0000],
    [0.1250, 0.3023, 0.1826, 0.3168, 1.0000, 0.0000, 0.0000], [0.5598, 0.4253, 0.4258, 0.3192, 0.0000, 1.0000, 0.0000], [0.5738, 0.7674, 0.6154, 0.4447, 0.0000, 0.0000, 1.0000],
    [0.5692, 0.8368, 0.5832, 0.4585, 0.0000, 0.0000, 1.0000], [0.4655, 0.7682, 0.3221, 0.2940, 0.0000, 1.0000, 0.0000], [0.5568, 0.7592, 0.6293, 0.5453, 0.0000, 1.0000, 0.0000],
    [0.8842, 0.7509, 0.5723, 0.5814, 0.0000, 0.0000, 1.0000], [0.7959, 0.9243, 0.7339, 0.7334, 0.0000, 0.0000, 1.0000], [0.7124, 0.7128, 0.6065, 0.6668, 0.0000, 0.0000, 1.0000],
    [0.6749, 0.8767, 0.6543, 0.7461, 0.0000, 0.0000, 1.0000], [0.3674, 0.4359, 0.4230, 0.2965, 1.0000, 0.0000, 0.0000], [0.3473, 0.0754, 0.2183, 0.1905, 1.0000, 0.0000, 0.0000],
    [0.6931, 0.5188, 0.5386, 0.5794, 0.0000, 1.0000, 0.0000], [0.6439, 0.4959, 0.4322, 0.4582, 0.0000, 1.0000, 0.0000], [0.5627, 0.4893, 0.6831, 0.5120, 0.0000, 1.0000, 0.0000],
    [0.5182, 0.7553, 0.6368, 0.4538, 0.0000, 1.0000, 0.0000], [0.6046, 0.7479, 0.6542, 0.4375, 0.0000, 1.0000, 0.0000], [0.6328, 0.6786, 0.7751, 0.6183, 0.0000, 0.0000, 1.0000],
    [0.3429, 0.4694, 0.2855, 0.2977, 1.0000, 0.0000, 0.0000], [0.6371, 0.5069, 0.5316, 0.4520, 0.0000, 1.0000, 0.0000], [0.6388, 0.6970, 0.6407, 0.7677, 0.0000, 0.0000, 1.0000],
    [0.3529, 0.5504, 0.3706, 0.4828, 0.0000, 1.0000, 0.0000], [0.4302, 0.3237, 0.6397, 0.4319, 0.0000, 1.0000, 0.0000], [0.7078, 0.9604, 0.7470, 0.6399, 0.0000, 0.0000, 1.0000],
    [0.7350, 0.8170, 0.7227, 0.6279, 0.0000, 0.0000, 1.0000], [0.7011, 0.2946, 0.6625, 0.4312, 0.0000, 1.0000, 0.0000], [0.5961, 0.3817, 0.6363, 0.3663, 0.0000, 1.0000, 0.0000],
    [0.0000, 0.2563, 0.2603, 0.3027, 1.0000, 0.0000, 0.0000], [0.5996, 0.5704, 0.6965, 0.6548, 0.0000, 0.0000, 1.0000], [0.4289, 0.3709, 0.3994, 0.3656, 0.0000, 1.0000, 0.0000],
    [0.2093, 0.3655, 0.3334, 0.1802, 1.0000, 0.0000, 0.0000], [0.2335, 0.2856, 0.3912, 0.1601, 1.0000, 0.0000, 0.0000], [0.3266, 0.7751, 0.4356, 0.3448, 0.0000, 1.0000, 0.0000],
    [0.2457, 0.1203, 0.1228, 0.2206, 1.0000, 0.0000, 0.0000], [0.4656, 0.4815, 0.4211, 0.4862, 0.0000, 1.0000, 0.0000], [0.7511, 0.8868, 0.5408, 0.6253, 0.0000, 0.0000, 1.0000],
    [0.7825, 0.9386, 0.6510, 0.6996, 0.0000, 0.0000, 1.0000], [0.3463, 0.4118, 0.2507, 0.0454, 1.0000, 0.0000, 0.0000], [0.5172, 0.1482, 0.3172, 0.2323, 1.0000, 0.0000, 0.0000],
    [0.6942, 0.4516, 0.5387, 0.5983, 0.0000, 1.0000, 0.0000], [0.7586, 0.7017, 0.7120, 0.7509, 0.0000, 0.0000, 1.0000], [0.6880, 0.6004, 0.6602, 0.4320, 0.0000, 1.0000, 0.0000],
    [0.4742, 0.5079, 0.4135, 0.4161, 0.0000, 1.0000, 0.0000], [0.4419, 0.5761, 0.4515, 0.4497, 0.0000, 1.0000, 0.0000], [0.3367, 0.4333, 0.2336, 0.1678, 1.0000, 0.0000, 0.0000],
    [0.4744, 0.4604, 0.1507, 0.4873, 1.0000, 0.0000, 0.0000], [0.7510, 0.4350, 0.5453, 0.4831, 0.0000, 1.0000, 0.0000], [0.4045, 0.5636, 0.2534, 0.5573, 0.0000, 1.0000, 0.0000],
    [0.1449, 0.1539, 0.2446, 0.0559, 1.0000, 0.0000, 0.0000], [0.3460, 0.2722, 0.1866, 0.5049, 1.0000, 0.0000, 0.0000], [0.2241, 0.2046, 0.3575, 0.2891, 1.0000, 0.0000, 0.0000],
    [0.1412, 0.2264, 0.4025, 0.2661, 1.0000, 0.0000, 0.0000], [0.5782, 0.6418, 0.7212, 0.6396, 0.0000, 0.0000, 1.0000], [0.9153, 0.6571, 0.8229, 0.6689, 0.0000, 0.0000, 1.0000],
    [0.6014, 0.7664, 0.6385, 0.5513, 0.0000, 0.0000, 1.0000], [0.7328, 0.8708, 0.8812, 0.7060, 0.0000, 0.0000, 1.0000], [0.4270, 0.6352, 0.6811, 0.3884, 0.0000, 1.0000, 0.0000],
    [0.6189, 0.1652, 0.4016, 0.3042, 1.0000, 0.0000, 0.0000], [0.2143, 0.3868, 0.1926, 0.0000, 1.0000, 0.0000, 0.0000], [0.5696, 0.7238, 0.7199, 0.6677, 0.0000, 0.0000, 1.0000],
    [0.8656, 0.6700, 0.6570, 0.6065, 0.0000, 0.0000, 1.0000], [0.9002, 0.6858, 0.7409, 0.7047, 0.0000, 0.0000, 1.0000], [0.4167, 0.5255, 0.5506, 0.4093, 0.0000, 1.0000, 0.0000],
    [0.8325, 0.4804, 0.7990, 0.7471, 0.0000, 0.0000, 1.0000], [0.4124, 0.1191, 0.4720, 0.3184, 1.0000, 0.0000, 0.0000], [1.0000, 1.0000, 0.7924, 0.7074, 0.0000, 0.0000, 1.0000],
    [0.5685, 0.6924, 0.6180, 0.5792, 0.0000, 1.0000, 0.0000], [0.6505, 0.4864, 0.2972, 0.4599, 0.0000, 1.0000, 0.0000], [0.8124, 0.7690, 0.9720, 1.0000, 0.0000, 0.0000, 1.0000],
    [0.9013, 0.7160, 1.0000, 0.8046, 0.0000, 0.0000, 1.0000], [0.8872, 0.7556, 0.9307, 0.6791, 0.0000, 0.0000, 1.0000], [0.3708, 0.2139, 0.2136, 0.4295, 1.0000, 0.0000, 0.0000],
    [0.5159, 0.4349, 0.3715, 0.4086, 0.0000, 1.0000, 0.0000], [0.6768, 0.6304, 0.8044, 0.4885, 0.0000, 0.0000, 1.0000], [0.1664, 0.2404, 0.2000, 0.3425, 1.0000, 0.0000, 0.0000],
    [0.2495, 0.2807, 0.4679, 0.2200, 1.0000, 0.0000, 0.0000], [0.2487, 0.2348, 0.0913, 0.1281, 1.0000, 0.0000, 0.0000], [0.5748, 0.8552, 0.5973, 0.7317, 0.0000, 0.0000, 1.0000],
    [0.3858, 0.7585, 0.3239, 0.3565, 0.0000, 1.0000, 0.0000], [0.3329, 0.4946, 0.5614, 0.3152, 0.0000, 1.0000, 0.0000], [0.3891, 0.4805, 0.7598, 0.4231, 0.0000, 1.0000, 0.0000],
    [0.2888, 0.4888, 0.1930, 0.0177, 1.0000, 0.0000, 0.0000], [0.3827, 0.4900, 0.2272, 0.3599, 0.0000, 1.0000, 0.0000], [0.6047, 0.4224, 0.6274, 0.5809, 0.0000, 1.0000, 0.0000],
    [0.9840, 0.7031, 0.6469, 0.4701, 0.0000, 0.0000, 1.0000], [0.6554, 0.6785, 0.9279, 0.7723, 0.0000, 0.0000, 1.0000], [0.0466, 0.3388, 0.0840, 0.0762, 1.0000, 0.0000, 0.0000],
    [0.6154, 0.8196, 0.6339, 0.7729, 0.0000, 0.0000, 1.0000], [0.8452, 0.8897, 0.8383, 0.6961, 0.0000, 0.0000, 1.0000], [0.6927, 0.7870, 0.7689, 0.7213, 0.0000, 0.0000, 1.0000],
    [0.4032, 0.6188, 0.4930, 0.5380, 0.0000, 1.0000, 0.0000], [0.4006, 0.3094, 0.3868, 0.0811, 1.0000, 0.0000, 0.0000], [0.7416, 0.7138, 0.6823, 0.6067, 0.0000, 0.0000, 1.0000],
    [0.7404, 0.6764, 0.8293, 0.4694, 0.0000, 0.0000, 1.0000], [0.7736, 0.7097, 0.6826, 0.8142, 0.0000, 0.0000, 1.0000], [0.5823, 0.9635, 0.3706, 0.5636, 0.0000, 1.0000, 0.0000],
    [0.2081, 0.3738, 0.3119, 0.3552, 1.0000, 0.0000, 0.0000], [0.5616, 0.8972, 0.5186, 0.6650, 0.0000, 0.0000, 1.0000], [0.6594, 0.8907, 0.6000, 0.7157, 0.0000, 0.0000, 1.0000],
    [0.3979, 0.3070, 0.3637, 0.1220, 1.0000, 0.0000, 0.0000], [0.2644, 0.0000, 0.3572, 0.1931, 1.0000, 0.0000, 0.0000], [0.4816, 0.4791, 0.4213, 0.5889, 0.0000, 1.0000, 0.0000],
    [0.0848, 0.0749, 0.4349, 0.3328, 1.0000, 0.0000, 0.0000], [0.4608, 0.6775, 0.3533, 0.3016, 0.0000, 1.0000, 0.0000], [0.4155, 0.6589, 0.5310, 0.5404, 0.0000, 1.0000, 0.0000],
    [0.3934, 0.6244, 0.4817, 0.4324, 0.0000, 1.0000, 0.0000], [0.5843, 0.8517, 0.8576, 0.7133, 0.0000, 0.0000, 1.0000], [0.1995, 0.3690, 0.3537, 0.3462, 1.0000, 0.0000, 0.0000],
    [0.3832, 0.2321, 0.0341, 0.2450, 1.0000, 0.0000, 0.0000],
])

# Subconjunto de VALIDACAO (18 amostras) - corresponde a Tabela 1 do enunciado
dados_validacao = np.array([
    [0.8622, 0.7101, 0.6236, 0.7894, 0.0000, 0.0000, 1.0000], [0.2741, 0.1552, 0.1333, 0.1516, 1.0000, 0.0000, 0.0000], [0.6772, 0.8516, 0.6543, 0.7573, 0.0000, 0.0000, 1.0000],
    [0.2178, 0.5039, 0.6415, 0.5039, 0.0000, 1.0000, 0.0000], [0.7260, 0.7500, 0.7007, 0.4953, 0.0000, 0.0000, 1.0000], [0.2473, 0.2941, 0.4248, 0.3087, 1.0000, 0.0000, 0.0000],
    [0.5682, 0.5683, 0.5054, 0.4426, 0.0000, 1.0000, 0.0000], [0.6566, 0.6715, 0.4952, 0.3951, 0.0000, 1.0000, 0.0000], [0.0705, 0.4717, 0.2921, 0.2954, 1.0000, 0.0000, 0.0000],
    [0.1187, 0.2568, 0.3140, 0.3037, 1.0000, 0.0000, 0.0000], [0.5673, 0.7011, 0.4083, 0.5552, 0.0000, 1.0000, 0.0000], [0.3164, 0.2251, 0.3526, 0.2560, 1.0000, 0.0000, 0.0000],
    [0.7884, 0.9568, 0.6825, 0.6398, 0.0000, 0.0000, 1.0000], [0.9633, 0.7850, 0.6777, 0.6059, 0.0000, 0.0000, 1.0000], [0.7739, 0.8505, 0.7934, 0.6626, 0.0000, 0.0000, 1.0000],
    [0.4219, 0.4136, 0.1408, 0.0940, 1.0000, 0.0000, 0.0000], [0.6616, 0.4365, 0.6597, 0.8129, 0.0000, 0.0000, 1.0000], [0.7325, 0.4761, 0.3888, 0.5683, 0.0000, 1.0000, 0.0000],
])

# Conjunto UNICO (validacao cruzada da etapa 3): treinamento + validacao = 148 amostras
dados_completos = np.vstack([dados_treinamento, dados_validacao])


# =============================================================================
# Funcoes de ativacao
# =============================================================================

def sigmoid(x):
    return 1 / (1 + np.exp(-x))

def sigmoid_deriv(x):
    return x * (1 - x)


# =============================================================================
# Rede PMC (Perceptron Multi-Camadas)
# =============================================================================

class MLP:

    def __init__(self, input_size, hidden_size, output_size, learning_rate=0.1, seed=None):
        if seed is not None:
            np.random.seed(seed)

        # Pesos e bias inicializados aleatoriamente entre 0 e 1
        self.W1 = np.random.rand(input_size, hidden_size)
        self.b1 = np.random.rand(hidden_size)
        self.W2 = np.random.rand(hidden_size, output_size)
        self.b2 = np.random.rand(output_size)

        self.learning_rate = learning_rate

    def forward(self, X):
        z1 = np.dot(X, self.W1) + self.b1
        a1 = sigmoid(z1)
        z2 = np.dot(a1, self.W2) + self.b2
        a2 = sigmoid(z2)
        return a2

    def eqm(self, X, y):
        """Erro Quadratico Medio na definicao do livro (Silva et al.):
        EQM = (1/p) * Sum_k [ (1/2) * Sum_j (d_j(k) - Y_j(k))^2 ]"""
        saidas = self.forward(X)
        return np.sum(0.5 * np.sum((y - saidas) ** 2, axis=1)) / len(X)

    def train(self, X, y, epsilon=1e-6, max_epochs=20000, min_epochs=200, patience=3):
        # min_epochs: ignora o criterio de parada nas primeiras epocas para nao
        #   parar prematuramente no plato inicial do gradiente por lote.
        # patience: o criterio |dEQM| <= epsilon precisa se manter por 'patience'
        #   epocas consecutivas, evitando parada em regioes de sela transitorias.
        eqm_list = []
        eqm_anterior = float('inf')
        consecutivas = 0

        for epoch in range(max_epochs):

            # ----- Forward (todo o conjunto de treino de uma vez) -----
            a1 = sigmoid(np.dot(X, self.W1) + self.b1)
            a2 = sigmoid(np.dot(a1, self.W2) + self.b2)

            # ----- Backward (gradiente acumulado sobre todas as amostras) -----
            erro = y - a2
            delta_output = erro * sigmoid_deriv(a2)
            error_hidden = np.dot(delta_output, self.W2.T)
            delta_hidden = error_hidden * sigmoid_deriv(a1)

            # ----- Atualizacao por LOTE (uma unica vez por epoca) -----
            self.W2 += self.learning_rate * np.dot(a1.T, delta_output)
            self.b2 += self.learning_rate * np.sum(delta_output, axis=0)
            self.W1 += self.learning_rate * np.dot(X.T, delta_hidden)
            self.b1 += self.learning_rate * np.sum(delta_hidden, axis=0)

            # EQM da epoca (definicao do livro) apos a atualizacao dos pesos
            eqm_atual = self.eqm(X, y)
            eqm_list.append(eqm_atual)

            if epoch % 200 == 0:
                print(f"    Epoca {epoch:5d} - EQM = {eqm_atual:.8f}")

            # Criterio de parada: variacao do EQM entre epocas <= epsilon,
            # mantida por 'patience' epocas consecutivas e somente apos min_epochs.
            variacao = abs(eqm_atual - eqm_anterior)
            if epoch >= min_epochs and variacao <= epsilon:
                consecutivas += 1
                if consecutivas >= patience:
                    break
            else:
                consecutivas = 0
            eqm_anterior = eqm_atual

        return eqm_list, epoch + 1

    def predict(self, X):
        return self.forward(X)


# =============================================================================
# Pos-processamento e metricas de avaliacao
# =============================================================================

def pos_processar(y_pred):
    """Criterio do enunciado: y_pos = 1 se y >= 0.5, senao 0."""
    return (y_pred >= 0.5).astype(int)


def confusion_matrix_3class(y_true, y_pred):
    """Matriz de confusao 3x3 baseada em argmax (one-hot -> classe)."""
    true_classes = np.argmax(y_true, axis=1)
    pred_classes = np.argmax(y_pred, axis=1)

    cm = np.zeros((3, 3), dtype=int)
    for t, p in zip(true_classes, pred_classes):
        cm[t][p] += 1
    return cm


def parametros_classificacao(cm):
    """A partir da matriz de confusao 3x3, calcula os parametros do enunciado:
    Nacertos, Nerros, Acuracia (global) e, para cada classe (abordagem
    um-contra-todos), Sensibilidade, Especificidade e Precisao, alem das
    medias (macro) entre as classes.

      TP = cm[i,i]                  (verdadeiros positivos da classe i)
      FP = soma(coluna i) - TP      (preditos como i, mas de outra classe)
      FN = soma(linha i)   - TP     (da classe i, mas preditos como outra)
      TN = total - TP - FP - FN     (nem preditos nem verdadeiros como i)

      Sensibilidade = TP / (TP + FN)
      Especificidade = TN / (TN + FP)
      Precisao       = TP / (TP + FP)
    """
    total = int(np.sum(cm))
    nacertos = int(np.trace(cm))
    nerros = total - nacertos
    acuracia = nacertos / total if total > 0 else 0.0

    sensibilidade, especificidade, precisao = [], [], []
    for i in range(cm.shape[0]):
        tp = cm[i, i]
        fp = np.sum(cm[:, i]) - tp
        fn = np.sum(cm[i, :]) - tp
        tn = total - tp - fp - fn
        sensibilidade.append(tp / (tp + fn) if (tp + fn) > 0 else 0.0)
        especificidade.append(tn / (tn + fp) if (tn + fp) > 0 else 0.0)
        precisao.append(tp / (tp + fp) if (tp + fp) > 0 else 0.0)

    return {
        "nacertos": nacertos,
        "nerros": nerros,
        "acuracia": acuracia,
        "sensibilidade": sensibilidade,
        "especificidade": especificidade,
        "precisao": precisao,
        "sensibilidade_media": float(np.mean(sensibilidade)),
        "especificidade_media": float(np.mean(especificidade)),
        "precisao_media": float(np.mean(precisao)),
    }


def contar_acertos(y_true, y_pos):
    """Conta amostras cujo vetor pos-processado bate exatamente com o desejado."""
    acertos = np.all(y_pos == y_true.astype(int), axis=1)
    return int(np.sum(acertos)), acertos


# =============================================================================
# Geracao da planilha Excel (Tabela 1 do enunciado)
# =============================================================================

def gerar_planilha_tabela1(X_val, y_val, y_pos, taxa_acertos, output_path):
    """Cria a planilha .xlsx no formato da Tabela 1 do PP04 e a preenche com
    as entradas, as saidas desejadas e as saidas pos-processadas da validacao."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela 1 - Validacao"

    # --- Estilos ---
    fonte_titulo  = Font(bold=True, size=12)
    fonte_cabec   = Font(bold=True, color="FFFFFF")
    preench_cabec = PatternFill("solid", fgColor="4472C4")
    preench_acerto = PatternFill("solid", fgColor="C6EFCE")  # verde claro
    preench_erro   = PatternFill("solid", fgColor="FFC7CE")  # vermelho claro
    centro = Alignment(horizontal="center", vertical="center")
    borda = Border(*(4 * (Side(style="thin", color="999999"),)))

    colunas = ["Amostra", "x1", "x2", "x3", "x4",
               "d1", "d2", "d3", "y1_pos", "y2_pos", "y3_pos"]

    # --- Titulo ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    titulo_cell = ws.cell(row=1, column=1, value="Tabela 1: Validacao da rede PMC da etapa 1")
    titulo_cell.font = fonte_titulo
    titulo_cell.alignment = centro

    # --- Cabecalho ---
    for j, nome in enumerate(colunas, start=1):
        c = ws.cell(row=2, column=j, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda

    # --- Linhas de dados ---
    _, acertos = contar_acertos(y_val, y_pos)
    for i in range(len(X_val)):
        linha = 3 + i
        valores = [i + 1,
                   *[round(float(v), 4) for v in X_val[i]],
                   *[int(v) for v in y_val[i]],
                   *[int(v) for v in y_pos[i]]]
        for j, val in enumerate(valores, start=1):
            c = ws.cell(row=linha, column=j, value=val)
            c.alignment = centro
            c.border = borda
        # destaca a amostra inteira conforme acerto/erro
        fill = preench_acerto if acertos[i] else preench_erro
        for j in range(9, 12):  # colunas das saidas pos-processadas
            ws.cell(row=linha, column=j).fill = fill

    # --- Linha de total de acertos ---
    linha_total = 3 + len(X_val)
    ws.merge_cells(start_row=linha_total, start_column=1,
                   end_row=linha_total, end_column=8)
    c = ws.cell(row=linha_total, column=1, value="Total de acertos (%)")
    c.font = Font(bold=True)
    c.alignment = centro
    c.border = borda
    ws.merge_cells(start_row=linha_total, start_column=9,
                   end_row=linha_total, end_column=11)
    c_val = ws.cell(row=linha_total, column=9, value=round(taxa_acertos, 2))
    c_val.font = Font(bold=True)
    c_val.alignment = centro
    c_val.border = borda

    # --- Largura das colunas ---
    ws.column_dimensions["A"].width = 10
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 9

    wb.save(output_path)


def gerar_planilha_etapa5(resultados, output_path):
    """Cria a planilha .xlsx com os resultados da etapa 5, no mesmo estilo da
    Tabela 1. Possui duas abas:
      - 'Parametros': uma linha por rede com Nacertos, Nerros, Acuracia,
        Sensibilidade, Especificidade e Precisao (por classe e media);
      - 'Matrizes de Confusao': a matriz de confusao 3x3 de cada rede.
    'resultados' e a lista retornada por etapa5_metricas (com 'nome', 'cm' e
    'parametros')."""
    classes = ["Tipo A", "Tipo B", "Tipo C"]

    # --- Estilos (mesma paleta da Tabela 1) ---
    fonte_titulo = Font(bold=True, size=12)
    fonte_cabec  = Font(bold=True, color="FFFFFF")
    fonte_bold   = Font(bold=True)
    preench_cabec = PatternFill("solid", fgColor="4472C4")  # azul
    preench_grupo = PatternFill("solid", fgColor="8EAADB")  # azul claro
    preench_diag  = PatternFill("solid", fgColor="C6EFCE")  # verde (acertos)
    preench_erro  = PatternFill("solid", fgColor="FFC7CE")  # vermelho (erros)
    centro = Alignment(horizontal="center", vertical="center")
    esq    = Alignment(horizontal="left", vertical="center")
    borda = Border(*(4 * (Side(style="thin", color="999999"),)))

    wb = Workbook()

    # =========================================================================
    # Aba 1 - Parametros de classificacao (uma linha por rede)
    # =========================================================================
    ws = wb.active
    ws.title = "Parametros"
    n_col = 16

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
    t = ws.cell(row=1, column=1,
                value="Etapa 5: Parametros de classificacao por rede")
    t.font = fonte_titulo
    t.alignment = centro

    # Cabecalho - colunas simples (mescladas verticalmente nas linhas 2 e 3)
    simples = [(1, "Rede"), (2, "Nacertos"), (3, "Nerros"), (4, "Acuracia (%)")]
    for col, nome in simples:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(row=2, column=col, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda

    # Cabecalho - grupos com 4 sub-colunas (Tipo A/B/C/Media)
    grupos = [("Sensibilidade (%)", 5), ("Especificidade (%)", 9),
              ("Precisao (%)", 13)]
    for nome, ini in grupos:
        ws.merge_cells(start_row=2, start_column=ini, end_row=2, end_column=ini + 3)
        c = ws.cell(row=2, column=ini, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda
        for k, sub in enumerate(["Tipo A", "Tipo B", "Tipo C", "Media"]):
            c2 = ws.cell(row=3, column=ini + k, value=sub)
            c2.font = fonte_bold
            c2.fill = preench_grupo
            c2.alignment = centro
            c2.border = borda

    # Linhas de dados (uma por rede)
    for idx, r in enumerate(resultados):
        linha = 4 + idx
        par = r["parametros"]
        valores = [
            r["nome"], par["nacertos"], par["nerros"],
            round(par["acuracia"] * 100, 2),
            *[round(v * 100, 2) for v in par["sensibilidade"]],
            round(par["sensibilidade_media"] * 100, 2),
            *[round(v * 100, 2) for v in par["especificidade"]],
            round(par["especificidade_media"] * 100, 2),
            *[round(v * 100, 2) for v in par["precisao"]],
            round(par["precisao_media"] * 100, 2),
        ]
        for j, val in enumerate(valores, start=1):
            c = ws.cell(row=linha, column=j, value=val)
            c.border = borda
            if j == 1:
                c.font = fonte_bold
                c.alignment = esq
            else:
                c.alignment = centro

    ws.column_dimensions["A"].width = 20
    for col in range(2, n_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # =========================================================================
    # Aba 2 - Matrizes de confusao (uma matriz 3x3 por rede, empilhadas)
    # =========================================================================
    ws2 = wb.create_sheet("Matrizes de Confusao")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t2 = ws2.cell(row=1, column=1,
                  value="Etapa 5: Matrizes de confusao "
                        "(linha = classe verdadeira, coluna = classe predita)")
    t2.font = fonte_titulo
    t2.alignment = centro

    linha = 3
    for r in resultados:
        cm = r["cm"]

        # Nome da rede
        ws2.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=5)
        c = ws2.cell(row=linha, column=1, value=r["nome"])
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        linha += 1

        # Cabecalho da matriz
        for j, nome in enumerate(["Real \\ Predito", "Tipo A", "Tipo B",
                                  "Tipo C", "Total"], start=1):
            cc = ws2.cell(row=linha, column=j, value=nome)
            cc.font = fonte_bold
            cc.fill = preench_grupo
            cc.alignment = centro
            cc.border = borda
        linha += 1

        # Conteudo da matriz (3 linhas)
        for i in range(3):
            rotulo = ws2.cell(row=linha, column=1, value=classes[i])
            rotulo.font = fonte_bold
            rotulo.fill = preench_grupo
            rotulo.alignment = centro
            rotulo.border = borda
            for j in range(3):
                cc = ws2.cell(row=linha, column=2 + j, value=int(cm[i, j]))
                cc.alignment = centro
                cc.border = borda
                if cm[i, j] > 0:
                    cc.fill = preench_diag if i == j else preench_erro
            tot = ws2.cell(row=linha, column=5, value=int(np.sum(cm[i])))
            tot.font = fonte_bold
            tot.alignment = centro
            tot.border = borda
            linha += 1
        linha += 1  # linha em branco separando as redes

    ws2.column_dimensions["A"].width = 18
    for col in ["B", "C", "D", "E"]:
        ws2.column_dimensions[col].width = 11

    wb.save(output_path)


# =============================================================================
# Graficos
# =============================================================================

def salvar_grafico_eqm(eqm_list, output_path, titulo):
    """Curva do EQM em funcao de cada epoca de treinamento (escala linear,
    convencao do livro-texto). Eixo x = epocas, eixo y = EQM."""
    epocas = np.arange(1, len(eqm_list) + 1)
    plt.figure(figsize=(7, 4.5))
    plt.plot(epocas, eqm_list, color='steelblue')
    plt.title(titulo)
    plt.xlabel("Epocas")
    plt.ylabel("EQM")
    plt.grid(True, linestyle=':', alpha=0.6)
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()


def salvar_grafico_eqm_comparativo(series, labels, output_path, titulo):
    """Sobrepoe varias curvas de EQM x epoca no mesmo grafico para comparacao
    (usado na etapa 4 com as 5 redes da etapa 3)."""
    plt.figure(figsize=(8, 5))
    for eqm_list, label in zip(series, labels):
        epocas = np.arange(1, len(eqm_list) + 1)
        plt.plot(epocas, eqm_list, label=label, linewidth=1.2)
    plt.title(titulo)
    plt.xlabel("Epocas")
    plt.ylabel("EQM")
    plt.grid(True, linestyle=':', alpha=0.6)
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()


def salvar_grafico_confusao(cm, output_path, titulo):
    """Plota a matriz de confusao como heatmap."""
    fig, ax = plt.subplots(figsize=(5, 4.5))
    im = ax.imshow(cm, interpolation='nearest', cmap=plt.cm.Blues)
    plt.colorbar(im, ax=ax)

    classes = ['Tipo A', 'Tipo B', 'Tipo C']
    tick_marks = np.arange(len(classes))
    ax.set_xticks(tick_marks)
    ax.set_xticklabels(classes, rotation=45)
    ax.set_yticks(tick_marks)
    ax.set_yticklabels(classes)

    thresh = cm.max() / 2.0
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], 'd'),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black",
                    fontsize=14)

    ax.set_ylabel('Classe verdadeira')
    ax.set_xlabel('Classe predita')
    ax.set_title(titulo)
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()


# =============================================================================
# Etapa 3 - Validacao cruzada (5 treinamentos com split aleatorio 80/20)
# =============================================================================

def etapa3_validacao_cruzada(dados, output_dir, n_redes=5, proporcao_treino=0.8):
    """Executa n_redes treinamentos. Em cada um, sorteia aleatoriamente
    proporcao_treino das amostras para treino e o restante para validacao.
    Sem semente fixa: cada execucao do programa gera divisoes diferentes."""
    n_total = len(dados)
    n_treino = int(round(proporcao_treino * n_total))
    n_val = n_total - n_treino

    print(f"  Conjunto unico: {n_total} amostras "
          f"({n_treino} treino / {n_val} validacao por rede)\n")

    resultados = []

    for i in range(n_redes):
        # Sorteio aleatorio das amostras (embaralha todos os indices e fatia)
        indices = np.random.permutation(n_total)
        idx_treino = indices[:n_treino]
        idx_val = indices[n_treino:]

        X_train = dados[idx_treino, :4]
        y_train = dados[idx_treino, 4:]
        X_val = dados[idx_val, :4]
        y_val = dados[idx_val, 4:]

        print(f"-- Rede {i+1} --")
        mlp = MLP(input_size=4, hidden_size=15, output_size=3, learning_rate=0.1)
        eqm_list, n_epocas = mlp.train(X_train, y_train, epsilon=1e-6, max_epochs=10000)

        # Validacao com pos-processamento
        y_val_pred = mlp.predict(X_val)
        y_val_pos = pos_processar(y_val_pred)
        n_acertos, _ = contar_acertos(y_val, y_val_pos)
        taxa = 100.0 * n_acertos / n_val

        print(f"   Epocas para convergir : {n_epocas}")
        print(f"   EQM final             : {eqm_list[-1]:.8f}")
        print(f"   Total de acertos      : {n_acertos}/{n_val} ({taxa:.2f}%)\n")

        resultados.append({
            "rede": i + 1,
            "epocas": n_epocas,
            "eqm_final": eqm_list[-1],
            "n_acertos": n_acertos,
            "n_val": n_val,
            "taxa": taxa,
            "eqm_list": eqm_list,
            "y_val": y_val,
            "y_val_pred": y_val_pred,
        })

    return resultados


# =============================================================================
# Etapa 4 - Graficos do EQM em funcao de cada epoca de treinamento
#           (para a rede da etapa 1 e para as 5 redes da etapa 3)
# =============================================================================

def etapa4_graficos_eqm(eqm_list_e1, resultados_e3, output_dir):
    """Traca o grafico do erro quadratico medio em funcao de cada epoca de
    treinamento para a rede da etapa 1 e para cada uma das redes da etapa 3,
    alem de um grafico comparativo sobrepondo as 5 redes da etapa 3."""
    gerados = []

    # --- Rede da etapa 1 ---
    caminho_e1 = os.path.join(output_dir, "eqm_etapa1.png")
    salvar_grafico_eqm(eqm_list_e1, caminho_e1,
                       "EQM x Epocas - Rede da Etapa 1")
    gerados.append(caminho_e1)

    # --- Cada rede da etapa 3 ---
    for r in resultados_e3:
        caminho = os.path.join(output_dir, f"eqm_etapa3_rede{r['rede']}.png")
        salvar_grafico_eqm(r["eqm_list"], caminho,
                           f"EQM x Epocas - Etapa 3 / Rede {r['rede']}")
        gerados.append(caminho)

    # --- Comparativo das 5 redes da etapa 3 ---
    caminho_comp = os.path.join(output_dir, "eqm_comparativo.png")
    salvar_grafico_eqm_comparativo(
        [r["eqm_list"] for r in resultados_e3],
        [f"Rede {r['rede']}" for r in resultados_e3],
        caminho_comp,
        "EQM x Epocas - Comparativo das redes da Etapa 3"
    )
    gerados.append(caminho_comp)

    for caminho in gerados:
        print(f"  Grafico salvo em: {caminho}")

    return gerados


# =============================================================================
# Etapa 5 - Matriz de confusao e parametros de classificacao por rede
#           (Nacertos, Nerros, Acuracia, Sensibilidade, Especificidade,
#            Precisao) para a rede da etapa 1 e para as 5 redes da etapa 3
# =============================================================================

def etapa5_metricas(redes, output_dir):
    """Para cada rede em 'redes' (dict com 'nome', 'slug', 'y_true', 'y_pred'),
    obtem a matriz de confusao 3x3 (classe = neuronio de maior saida, argmax),
    calcula os parametros de classificacao e salva o heatmap da matriz."""
    classes = ['Tipo A', 'Tipo B', 'Tipo C']
    resultados = []

    for rede in redes:
        cm = confusion_matrix_3class(rede["y_true"], rede["y_pred"])
        par = parametros_classificacao(cm)

        print(f"\n>> {rede['nome']}")
        print("   Matriz de confusao (linha = verdadeiro, coluna = predito):")
        print("            " + "".join(f"{c:>8}" for c in classes))
        for i, c in enumerate(classes):
            print(f"   {c:>8} " + "".join(f"{cm[i, j]:>8d}" for j in range(3)))

        print(f"   Nacertos   : {par['nacertos']}")
        print(f"   Nerros     : {par['nerros']}")
        print(f"   Acuracia   : {par['acuracia'] * 100:.2f}%")
        print(f"   {'Parametro':<14}{'Tipo A':>9}{'Tipo B':>9}{'Tipo C':>9}{'Media':>10}")
        for nome, chave in [("Sensibilidade", "sensibilidade"),
                            ("Especificidade", "especificidade"),
                            ("Precisao", "precisao")]:
            vals = "".join(f"{v * 100:8.2f}%" for v in par[chave])
            media = par[f"{chave}_media"] * 100
            print(f"   {nome:<14}{vals}{media:9.2f}%")

        caminho = os.path.join(output_dir, f"confusao_{rede['slug']}.png")
        salvar_grafico_confusao(cm, caminho,
                                f"Matriz de Confusao - {rede['nome']}")
        print(f"   Grafico salvo em: {caminho}")

        resultados.append({**rede, "cm": cm, "parametros": par})

    return resultados


# =============================================================================
# Main - Etapa 1: Treinamento da rede PMC
# =============================================================================

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "graphics")
    os.makedirs(output_dir, exist_ok=True)

    # Separa entradas (x1..x4) e saidas desejadas (d1, d2, d3)
    X_train = dados_treinamento[:, :4]
    y_train = dados_treinamento[:, 4:]
    X_val   = dados_validacao[:, :4]
    y_val   = dados_validacao[:, 4:]

    print(f"Amostras de treinamento : {X_train.shape[0]}")
    print(f"Amostras de validacao   : {X_val.shape[0]}")
    print(f"Entradas: {X_train.shape[1]}  |  Saidas: {y_train.shape[1]}\n")

    # -------------------------------------------------------------------------
    # ETAPA 1 - Treinamento da rede PMC
    # Topologia: 4 entradas - 15 neuronios ocultos - 3 saidas
    # Funcao de ativacao logistica, eta = 0.1, epsilon = 1e-6
    # -------------------------------------------------------------------------
    print("=" * 60)
    print("ETAPA 1 - TREINAMENTO DA REDE PMC")
    print("=" * 60)

    mlp = MLP(input_size=4, hidden_size=15, output_size=3, learning_rate=0.1)
    eqm_list, n_epocas = mlp.train(X_train, y_train, epsilon=1e-6, max_epochs=10000)

    print(f"\n  Treinamento concluido.")
    print(f"  Epocas             : {n_epocas}")
    print(f"  EQM inicial        : {eqm_list[0]:.8f}")
    print(f"  EQM final          : {eqm_list[-1]:.8f}")

    # Guarda a curva de EQM da etapa 1 para tracar o grafico na etapa 4
    eqm_list_e1 = eqm_list

    # Verificacao rapida nos proprios dados de treino (sanidade)
    y_train_pos = pos_processar(mlp.predict(X_train))
    acertos_treino = np.sum(np.all(y_train_pos == y_train.astype(int), axis=1))
    print(f"  Acertos no treino  : {acertos_treino}/{len(X_train)} "
          f"({100.0 * acertos_treino / len(X_train):.2f}%)")

    # -------------------------------------------------------------------------
    # ETAPA 2 - Validacao da rede e preenchimento da Tabela 1
    # Pos-processamento: y_pos = 1 se y >= 0.5, senao 0
    # -------------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("ETAPA 2 - VALIDACAO DA REDE (TABELA 1)")
    print("=" * 60)

    y_val_pred = mlp.predict(X_val)
    y_val_pos  = pos_processar(y_val_pred)

    n_acertos, acertos = contar_acertos(y_val, y_val_pos)
    taxa_acertos = 100.0 * n_acertos / len(X_val)

    # Impressao da tabela no terminal
    cabecalho = (f"{'Am':>3} | {'x1':>6} {'x2':>6} {'x3':>6} {'x4':>6} | "
                 f"{'d1':>2} {'d2':>2} {'d3':>2} | "
                 f"{'y1':>2} {'y2':>2} {'y3':>2} | OK")
    print(cabecalho)
    print("-" * len(cabecalho))
    for i in range(len(X_val)):
        ok = "sim" if acertos[i] else "NAO"
        print(f"{i+1:>3} | "
              f"{X_val[i,0]:6.4f} {X_val[i,1]:6.4f} {X_val[i,2]:6.4f} {X_val[i,3]:6.4f} | "
              f"{int(y_val[i,0]):>2} {int(y_val[i,1]):>2} {int(y_val[i,2]):>2} | "
              f"{y_val_pos[i,0]:>2} {y_val_pos[i,1]:>2} {y_val_pos[i,2]:>2} | {ok}")

    print(f"\n  Total de acertos: {n_acertos}/{len(X_val)} = {taxa_acertos:.2f}%")

    # Geracao da planilha Excel no formato da Tabela 1
    planilha_path = os.path.join(base_dir, "Tabela1_validacao.xlsx")
    gerar_planilha_tabela1(X_val, y_val, y_val_pos, taxa_acertos, planilha_path)
    print(f"  Planilha Excel salva em: {planilha_path}")

    # -------------------------------------------------------------------------
    # ETAPA 3 - Validacao cruzada: 5 novos treinamentos com split aleatorio 80/20
    # -------------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("ETAPA 3 - VALIDACAO CRUZADA (5 TREINAMENTOS - SPLIT 80/20)")
    print("=" * 60)

    resultados_e3 = etapa3_validacao_cruzada(dados_completos, output_dir,
                                             n_redes=5, proporcao_treino=0.8)

    print("-" * 60)
    print("RESUMO DA ETAPA 3")
    print("-" * 60)
    print(f"{'Rede':>5} | {'Epocas':>7} | {'EQM final':>12} | {'Acertos':>10} | {'Taxa':>7}")
    print("-" * 60)
    for r in resultados_e3:
        print(f"{r['rede']:>5} | {r['epocas']:>7} | {r['eqm_final']:>12.8f} | "
              f"{r['n_acertos']:>3}/{r['n_val']:<6} | {r['taxa']:>6.2f}%")

    # -------------------------------------------------------------------------
    # ETAPA 4 - Grafico do EQM em funcao de cada epoca de treinamento
    #           para as redes das etapas 1 e 3
    # -------------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("ETAPA 4 - GRAFICOS DO EQM x EPOCAS (ETAPAS 1 E 3)")
    print("=" * 60)

    etapa4_graficos_eqm(eqm_list_e1, resultados_e3, output_dir)

    # -------------------------------------------------------------------------
    # ETAPA 5 - Matriz de confusao e parametros de classificacao por rede
    # -------------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("ETAPA 5 - MATRIZ DE CONFUSAO E PARAMETROS POR REDE")
    print("=" * 60)

    # Monta a lista de todas as redes desenvolvidas: a da etapa 1 e as 5 da
    # etapa 3, cada uma avaliada no seu respectivo conjunto de validacao.
    redes_avaliacao = [
        {"nome": "Rede da Etapa 1", "slug": "etapa1",
         "y_true": y_val, "y_pred": y_val_pred},
    ]
    for r in resultados_e3:
        redes_avaliacao.append({
            "nome": f"Etapa 3 / Rede {r['rede']}",
            "slug": f"etapa3_rede{r['rede']}",
            "y_true": r["y_val"],
            "y_pred": r["y_val_pred"],
        })

    resultados_e5 = etapa5_metricas(redes_avaliacao, output_dir)

    # Planilha Excel com os parametros e as matrizes de confusao (etapa 5)
    planilha_e5_path = os.path.join(base_dir, "Etapa5_metricas.xlsx")
    gerar_planilha_etapa5(resultados_e5, planilha_e5_path)
    print(f"\n  Planilha Excel da etapa 5 salva em: {planilha_e5_path}")


if __name__ == "__main__":
    main()
